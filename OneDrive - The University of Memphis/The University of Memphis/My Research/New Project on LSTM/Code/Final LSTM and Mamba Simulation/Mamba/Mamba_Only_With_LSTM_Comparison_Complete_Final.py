"""
MAMBA-ONLY ACTUATOR SYSTEM IDENTIFICATION
=========================================

Purpose
-------
Train a Mamba-style selective state-space model on the same actuator data and
split used by the LSTM simulation, so the results can be compared fairly.

Data use
--------
Development: 67, 87, 107, and 127 mA chirps
Independent test: 147 mA chirp

Outputs
-------
1. Mamba one-step prediction using measured displacement/force history.
2. Mamba free-running rollout using its own predicted output history.
3. Persistence baseline.
4. Metrics, prediction CSV files, model weights, normalizer parameters,
   training curves, tracking plots, error plots, parity/regression plots,
   a free-running stability diagnostic, and a complete comparison with the
   newest LSTM result set found under 03_Complete_Results. The comparison
   includes common metrics, tracking, errors, regression, free-running
   stability, and an automatically generated model-selection summary.

Important implementation note
-----------------------------
This is a portable pure-PyTorch implementation of the Mamba-1 selective SSM
core. It follows the input-dependent Delta, B, and C recurrence used by Mamba,
with causal depthwise convolution, gating, residual connections, and RMSNorm.
It does NOT use the official fused CUDA selective-scan kernel, so it runs on
Windows/CPU but is slower than the official Linux/CUDA implementation.

How to run
----------
Place this file in the same folder as:
    COMSOL_07_13_2026.xlsx

Then run it directly in VS Code.

Quick test mode in PowerShell:
    $env:QUICK_MODE="1"
    python Mamba_Only_Actuator_System_Identification.py

Final run:
    Remove QUICK_MODE or set it to 0, then run again.
"""

from __future__ import annotations

# -----------------------------------------------------------------------------
# 0. PACKAGE CHECK
# -----------------------------------------------------------------------------

import importlib.util
import subprocess
import sys

REQUIRED_PACKAGES = {
    "numpy": "numpy",
    "matplotlib": "matplotlib",
    "openpyxl": "openpyxl",
    "torch": "torch",
}

missing_packages = [
    pip_name
    for import_name, pip_name in REQUIRED_PACKAGES.items()
    if importlib.util.find_spec(import_name) is None
]

if missing_packages:
    print("Installing missing packages:", ", ".join(missing_packages))
    subprocess.check_call([sys.executable, "-m", "pip", "install", *missing_packages])

# -----------------------------------------------------------------------------
# 1. IMPORTS AND SETTINGS
# -----------------------------------------------------------------------------

import copy
import csv
import json
import math
import os
import random
import re
import shutil
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import torch
import torch.nn.functional as F
from torch import nn
from torch.utils.data import DataLoader, Dataset, Subset

# Same temporal window as the current LSTM experiment.
WINDOW = 120

# Small Mamba architecture suitable for the available data size.
D_MODEL = 24
D_STATE = 6
D_CONV = 4
EXPAND = 2
N_MAMBA_BLOCKS = 1
DROPOUT = 0.03
DELTA_OUTPUT_SCALE = 0.1

ONE_STEP_EPOCHS = 6
ROLLOUT_EPOCHS = 2
ROLLOUT_HORIZON = 50

ONE_STEP_BATCH_SIZE = 256
ROLLOUT_BATCH_SIZE = 32
LEARNING_RATE = 8e-4
ROLLOUT_LEARNING_RATE = 1e-4
WEIGHT_DECAY = 1e-6

TRAIN_STRIDE = 20
VALIDATION_STRIDE = 10
ROLLOUT_STRIDE = 100

CYCLIC_TRAIN_SAMPLES = 1000
CYCLIC_VALIDATION_SAMPLES = 500
CYCLIC_DEVELOPMENT_TEST_SAMPLES = 500
GUARD_GAP_SAMPLES = WINDOW

RANDOM_SEED = 123
AUTO_GIT_PUSH = True
GIT_REPOSITORY_SSH = "git@github.com:hzolfaghari2022/LSTM_Modelling.git"
GIT_USER_NAME = "Hussein Zolfaghari"
GIT_USER_EMAIL = "h.zolfaghari2015@gmail.com"

# LSTM comparison settings. Run the LSTM simulation at least once before this
# Mamba script. By default, the newest compatible LSTM result folder is used.
# To force a specific folder in PowerShell, use for example:
#   $env:LSTM_RESULTS_FOLDER="C:\\path\\to\\03_Complete_Results\\20260724_120000"
LSTM_RESULTS_FOLDER = os.environ.get("LSTM_RESULTS_FOLDER", "").strip()
REQUIRE_LSTM_COMPARISON = True
COMPARISON_SCATTER_MAX_POINTS = 5000

DEVELOPMENT_SHEETS = [
    "DC_Offset_67mA",
    "DC_Offset_87mA",
    "DC_Offset_107mA",
    "DC_Offset_127mA",
]
TEST_SHEET = "DC_Offset_147mA"

QUICK_MODE = os.environ.get("QUICK_MODE", "0") == "1"
ROOT = Path(__file__).resolve().parent

torch.set_num_threads(max(1, min(4, os.cpu_count() or 1)))

# -----------------------------------------------------------------------------
# 2. DATA STRUCTURES
# -----------------------------------------------------------------------------

@dataclass
class Block:
    sheet: str
    start: int
    end: int
    role: str


@dataclass
class Normalizer:
    input_mean: np.ndarray
    input_std: np.ndarray
    output_mean: np.ndarray
    output_std: np.ndarray

    def normalize_input(self, values: np.ndarray) -> np.ndarray:
        return (values - self.input_mean) / self.input_std

    def normalize_output(self, values: np.ndarray) -> np.ndarray:
        return (values - self.output_mean) / self.output_std

    def restore_output(self, values: np.ndarray) -> np.ndarray:
        return values * self.output_std + self.output_mean

# -----------------------------------------------------------------------------
# 3. REPRODUCIBILITY AND EXCEL READING
# -----------------------------------------------------------------------------

def set_random_seed() -> None:
    random.seed(RANDOM_SEED)
    np.random.seed(RANDOM_SEED)
    torch.manual_seed(RANDOM_SEED)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(RANDOM_SEED)


def find_workbook() -> Path:
    preferred_names = [
        "COMSOL_07_13_2026.xlsx",
        "COMSOL_07_13_2026(1).xlsx",
    ]
    for name in preferred_names:
        candidate = ROOT / name
        if candidate.exists():
            return candidate

    workbooks = [p for p in ROOT.glob("*.xlsx") if not p.name.startswith("~$")]
    if len(workbooks) == 1:
        return workbooks[0]

    raise FileNotFoundError(
        "Place COMSOL_07_13_2026.xlsx in the same folder as this Python file."
    )


def open_workbook_safely(workbook_path: Path) -> openpyxl.Workbook:
    temporary_folder = Path(tempfile.gettempdir()) / "Mamba_COMSOL_Data"
    temporary_folder.mkdir(parents=True, exist_ok=True)
    temporary_copy = temporary_folder / f"COMSOL_copy_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    for attempt in range(1, 4):
        try:
            shutil.copyfile(workbook_path, temporary_copy)
            print(f"Using temporary workbook copy: {temporary_copy}")
            return openpyxl.load_workbook(
                temporary_copy,
                read_only=True,
                data_only=True,
            )
        except PermissionError:
            if attempt == 3:
                raise PermissionError(
                    "The Excel workbook is locked. Close Excel and the File Explorer preview pane."
                )
            input("Close the workbook, then press Enter to retry...")

    raise RuntimeError("Could not open the workbook.")


def load_sheet(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    maximum_rows: int | None = None,
) -> np.ndarray:
    """Read time, displacement, coil current, and Lorentz force."""
    if sheet_name not in workbook.sheetnames:
        raise KeyError(
            f"Missing sheet {sheet_name}. Available sheets: {workbook.sheetnames}"
        )

    rows: list[list[float]] = []
    worksheet = workbook[sheet_name]

    for row in worksheet.iter_rows(min_col=1, max_col=4, values_only=True):
        try:
            values = [float(row[i]) for i in range(4)]
        except (TypeError, ValueError):
            continue

        if np.all(np.isfinite(values)):
            rows.append(values)
            if maximum_rows is not None and len(rows) >= maximum_rows:
                break

    data = np.asarray(rows, dtype=np.float32)
    if len(data) <= WINDOW + 10:
        raise ValueError(f"Not enough numeric data in {sheet_name}")

    original_length = len(data)
    _, unique_indices = np.unique(data[:, 0], return_index=True)
    data = data[np.sort(unique_indices)]
    removed = original_length - len(data)
    if removed:
        print(f"{sheet_name}: removed {removed} duplicate time rows.")

    order = np.argsort(data[:, 0])
    data = data[order]

    dt = np.diff(data[:, 0])
    median_dt = float(np.median(dt))
    if not np.allclose(dt, median_dt, rtol=1e-3, atol=1e-9):
        print(f"Warning: {sheet_name} has nonuniform time steps; median dt={median_dt:.6g} s")

    return data


def dc_offset_from_sheet(sheet_name: str) -> float:
    match = re.search(r"_(\d+)mA", sheet_name)
    if match is None:
        raise ValueError(f"Cannot infer DC offset from sheet name: {sheet_name}")
    return float(match.group(1)) / 1000.0


def known_input_features(raw: np.ndarray, sheet_name: str) -> np.ndarray:
    current = raw[:, 2]
    delta_current = np.r_[0.0, np.diff(current)].astype(np.float32)
    dc_offset = np.full_like(current, dc_offset_from_sheet(sheet_name))
    return np.column_stack([current, delta_current, dc_offset]).astype(np.float32)


def measured_outputs(raw: np.ndarray) -> np.ndarray:
    return raw[:, [1, 3]].astype(np.float32)

# -----------------------------------------------------------------------------
# 4. CYCLIC DATA SPLIT AND NORMALIZATION
# -----------------------------------------------------------------------------

def create_cyclic_blocks(
    development_data: dict[str, np.ndarray],
    maximum_horizon: int,
) -> list[Block]:
    blocks: list[Block] = []
    cycle = [
        ("training", CYCLIC_TRAIN_SAMPLES),
        ("gap", GUARD_GAP_SAMPLES),
        ("validation", CYCLIC_VALIDATION_SAMPLES),
        ("gap", GUARD_GAP_SAMPLES),
        ("development_test", CYCLIC_DEVELOPMENT_TEST_SAMPLES),
        ("gap", GUARD_GAP_SAMPLES),
    ]
    minimum_useful_length = WINDOW + maximum_horizon + 1

    for sheet_name in DEVELOPMENT_SHEETS:
        n_samples = len(development_data[sheet_name])
        start = 0
        while start < n_samples:
            for role, length in cycle:
                end = min(start + length, n_samples)
                if end <= start:
                    break
                if role == "gap" or end - start >= minimum_useful_length:
                    blocks.append(Block(sheet_name, start, end, role))
                start = end
                if n_samples - start < minimum_useful_length:
                    if start < n_samples:
                        blocks.append(Block(sheet_name, start, n_samples, "gap"))
                    start = n_samples
                    break

    return blocks


def create_normalizer(
    development_data: dict[str, np.ndarray],
    blocks: list[Block],
) -> Normalizer:
    input_rows: list[np.ndarray] = []
    output_rows: list[np.ndarray] = []

    for block in blocks:
        if block.role != "training":
            continue
        input_rows.append(
            known_input_features(development_data[block.sheet], block.sheet)[block.start:block.end]
        )
        output_rows.append(measured_outputs(development_data[block.sheet])[block.start:block.end])

    all_inputs = np.concatenate(input_rows)
    all_outputs = np.concatenate(output_rows)

    input_std = all_inputs.std(axis=0)
    output_std = all_outputs.std(axis=0)
    input_std[input_std < 1e-12] = 1.0
    output_std[output_std < 1e-12] = 1.0

    return Normalizer(
        input_mean=all_inputs.mean(axis=0).astype(np.float32),
        input_std=input_std.astype(np.float32),
        output_mean=all_outputs.mean(axis=0).astype(np.float32),
        output_std=output_std.astype(np.float32),
    )


def normalize_all_records(
    raw_data: dict[str, np.ndarray],
    normalizer: Normalizer,
) -> tuple[dict[str, np.ndarray], dict[str, np.ndarray]]:
    input_data: dict[str, np.ndarray] = {}
    output_data: dict[str, np.ndarray] = {}

    for sheet_name, raw in raw_data.items():
        input_data[sheet_name] = normalizer.normalize_input(
            known_input_features(raw, sheet_name)
        ).astype(np.float32)
        output_data[sheet_name] = normalizer.normalize_output(
            measured_outputs(raw)
        ).astype(np.float32)

    return input_data, output_data

# -----------------------------------------------------------------------------
# 5. DATASETS
# -----------------------------------------------------------------------------

class OneStepDataset(Dataset):
    """Measured input/output history -> next measured output."""

    def __init__(
        self,
        input_data: dict[str, np.ndarray],
        output_data: dict[str, np.ndarray],
        blocks: list[Block],
        role: str,
        stride: int,
    ) -> None:
        self.input_data = input_data
        self.output_data = output_data
        self.indices: list[tuple[str, int]] = []

        for block in blocks:
            if block.role != role:
                continue
            for target_index in range(block.start + WINDOW, block.end, stride):
                self.indices.append((block.sheet, target_index))

    def __len__(self) -> int:
        return len(self.indices)

    def __getitem__(self, index: int):
        sheet_name, target_index = self.indices[index]
        start = target_index - WINDOW
        return (
            torch.from_numpy(self.input_data[sheet_name][start:target_index]),
            torch.from_numpy(self.output_data[sheet_name][start:target_index]),
            torch.from_numpy(self.output_data[sheet_name][target_index]),
        )


class RolloutDataset(Dataset):
    """Warm-up history followed by a recursive prediction horizon."""

    def __init__(
        self,
        input_data: dict[str, np.ndarray],
        output_data: dict[str, np.ndarray],
        blocks: list[Block],
        role: str,
        stride: int,
        horizon: int,
    ) -> None:
        self.input_data = input_data
        self.output_data = output_data
        self.horizon = horizon
        self.indices: list[tuple[str, int]] = []

        for block in blocks:
            if block.role != role:
                continue
            last_start = block.end - WINDOW - horizon
            for start in range(block.start, max(block.start, last_start + 1), stride):
                if start + WINDOW + horizon <= block.end:
                    self.indices.append((block.sheet, start))

    def __len__(self) -> int:
        return len(self.indices)

    def __getitem__(self, index: int):
        sheet_name, start = self.indices[index]
        end = start + WINDOW + self.horizon
        return (
            torch.from_numpy(self.input_data[sheet_name][start:end]),
            torch.from_numpy(self.output_data[sheet_name][start:end]),
        )

# -----------------------------------------------------------------------------
# 6. PURE-PYTORCH MAMBA-1 BLOCK
# -----------------------------------------------------------------------------

class RMSNorm(nn.Module):
    def __init__(self, dimension: int, epsilon: float = 1e-5) -> None:
        super().__init__()
        self.weight = nn.Parameter(torch.ones(dimension))
        self.epsilon = epsilon

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        rms = torch.rsqrt(x.pow(2).mean(dim=-1, keepdim=True) + self.epsilon)
        return x * rms * self.weight


def inverse_softplus(x: torch.Tensor) -> torch.Tensor:
    return x + torch.log(-torch.expm1(-x))


class SelectiveMambaMixer(nn.Module):
    """
    Portable recurrent Mamba-1 selective SSM mixer.

    The step function keeps a causal-convolution state and an SSM state, so
    free-running prediction advances in constant work per new sample instead
    of reprocessing the full 120-sample window at every rollout step.
    """

    def __init__(
        self,
        d_model: int,
        d_state: int,
        d_conv: int,
        expand: int,
    ) -> None:
        super().__init__()
        self.d_model = d_model
        self.d_state = d_state
        self.d_conv = d_conv
        self.d_inner = d_model * expand
        self.dt_rank = max(1, math.ceil(d_model / 16))

        self.in_proj = nn.Linear(d_model, 2 * self.d_inner, bias=False)
        self.depthwise_conv = nn.Conv1d(
            self.d_inner,
            self.d_inner,
            kernel_size=d_conv,
            groups=self.d_inner,
            bias=True,
        )
        self.x_proj = nn.Linear(
            self.d_inner,
            self.dt_rank + 2 * d_state,
            bias=False,
        )
        self.dt_proj = nn.Linear(self.dt_rank, self.d_inner, bias=True)

        A = torch.arange(1, d_state + 1, dtype=torch.float32).repeat(self.d_inner, 1)
        self.A_log = nn.Parameter(torch.log(A))
        self.D = nn.Parameter(torch.ones(self.d_inner))

        dt = torch.exp(
            torch.rand(self.d_inner) * (math.log(0.1) - math.log(0.001))
            + math.log(0.001)
        ).clamp(min=1e-4)
        with torch.no_grad():
            self.dt_proj.bias.copy_(inverse_softplus(dt))

        self.out_proj = nn.Linear(self.d_inner, d_model, bias=False)

    def initial_state(
        self,
        batch_size: int,
        device: torch.device,
        dtype: torch.dtype,
    ) -> tuple[torch.Tensor, torch.Tensor]:
        conv_state = torch.zeros(
            batch_size,
            self.d_inner,
            max(self.d_conv - 1, 0),
            device=device,
            dtype=dtype,
        )
        ssm_state = torch.zeros(
            batch_size,
            self.d_inner,
            self.d_state,
            device=device,
            dtype=dtype,
        )
        return conv_state, ssm_state

    def step(
        self,
        hidden_t: torch.Tensor,
        conv_state: torch.Tensor,
        ssm_state: torch.Tensor,
    ) -> tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
        xz = self.in_proj(hidden_t)
        x_t, z_t = xz.chunk(2, dim=-1)

        if self.d_conv > 1:
            conv_window = torch.cat([conv_state, x_t.unsqueeze(-1)], dim=-1)
            new_conv_state = conv_window[:, :, 1:]
        else:
            conv_window = x_t.unsqueeze(-1)
            new_conv_state = conv_state

        kernel = self.depthwise_conv.weight[:, 0, :].unsqueeze(0)
        conv_output = torch.sum(conv_window * kernel, dim=-1)
        if self.depthwise_conv.bias is not None:
            conv_output = conv_output + self.depthwise_conv.bias
        x_t = F.silu(conv_output)

        projected = self.x_proj(x_t)
        dt_raw, B_t, C_t = torch.split(
            projected,
            [self.dt_rank, self.d_state, self.d_state],
            dim=-1,
        )
        delta_t = F.softplus(self.dt_proj(dt_raw)).clamp(max=1.0)

        A = -torch.exp(self.A_log.float()).to(dtype=x_t.dtype, device=x_t.device)
        D = self.D.to(dtype=x_t.dtype, device=x_t.device)

        dA = torch.exp(delta_t.unsqueeze(-1) * A.unsqueeze(0))
        dB = delta_t.unsqueeze(-1) * B_t.unsqueeze(1)
        new_ssm_state = ssm_state * dA + x_t.unsqueeze(-1) * dB

        y_t = torch.sum(new_ssm_state * C_t.unsqueeze(1), dim=-1) + D * x_t
        y_t = y_t * F.silu(z_t)
        return self.out_proj(y_t), new_conv_state, new_ssm_state

    def forward(self, hidden_states: torch.Tensor) -> torch.Tensor:
        batch_size, length, _ = hidden_states.shape
        conv_state, ssm_state = self.initial_state(
            batch_size,
            hidden_states.device,
            hidden_states.dtype,
        )
        outputs: list[torch.Tensor] = []
        for step_index in range(length):
            output_t, conv_state, ssm_state = self.step(
                hidden_states[:, step_index, :],
                conv_state,
                ssm_state,
            )
            outputs.append(output_t)
        return torch.stack(outputs, dim=1)


class MambaBlock(nn.Module):
    def __init__(
        self,
        d_model: int,
        d_state: int,
        d_conv: int,
        expand: int,
        dropout: float,
    ) -> None:
        super().__init__()
        self.norm = RMSNorm(d_model)
        self.mixer = SelectiveMambaMixer(d_model, d_state, d_conv, expand)
        self.dropout = nn.Dropout(dropout)

    def initial_state(
        self,
        batch_size: int,
        device: torch.device,
        dtype: torch.dtype,
    ) -> tuple[torch.Tensor, torch.Tensor]:
        return self.mixer.initial_state(batch_size, device, dtype)

    def step(
        self,
        hidden_t: torch.Tensor,
        state: tuple[torch.Tensor, torch.Tensor],
    ) -> tuple[torch.Tensor, tuple[torch.Tensor, torch.Tensor]]:
        conv_state, ssm_state = state
        mixed, conv_state, ssm_state = self.mixer.step(
            self.norm(hidden_t),
            conv_state,
            ssm_state,
        )
        return hidden_t + self.dropout(mixed), (conv_state, ssm_state)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        return x + self.dropout(self.mixer(self.norm(x)))


class MambaActuatorModel(nn.Module):
    """Mamba-only measured-feedback model for displacement and force."""

    def __init__(self) -> None:
        super().__init__()
        self.input_projection = nn.Linear(5, D_MODEL)
        self.blocks = nn.ModuleList(
            [
                MambaBlock(D_MODEL, D_STATE, D_CONV, EXPAND, DROPOUT)
                for _ in range(N_MAMBA_BLOCKS)
            ]
        )
        self.final_norm = RMSNorm(D_MODEL)
        self.head = nn.Sequential(
            nn.Linear(D_MODEL, 32),
            nn.SiLU(),
            nn.Linear(32, 2),
        )
        # Start close to the persistence model: next output ≈ previous output.
        nn.init.zeros_(self.head[-1].weight)
        nn.init.zeros_(self.head[-1].bias)

    def initial_state(
        self,
        batch_size: int,
        device: torch.device,
        dtype: torch.dtype,
    ) -> list[tuple[torch.Tensor, torch.Tensor]]:
        return [
            block.initial_state(batch_size, device, dtype)
            for block in self.blocks
        ]

    def step_token(
        self,
        token: torch.Tensor,
        states: list[tuple[torch.Tensor, torch.Tensor]],
    ) -> tuple[torch.Tensor, list[tuple[torch.Tensor, torch.Tensor]]]:
        hidden = self.input_projection(token)
        new_states: list[tuple[torch.Tensor, torch.Tensor]] = []
        for block, state in zip(self.blocks, states):
            hidden, new_state = block.step(hidden, state)
            new_states.append(new_state)
        hidden = self.final_norm(hidden)
        return hidden, new_states

    def warmup(
        self,
        input_history: torch.Tensor,
        output_history: torch.Tensor,
    ) -> tuple[torch.Tensor, list[tuple[torch.Tensor, torch.Tensor]]]:
        tokens = torch.cat([input_history, output_history], dim=-1)
        states = self.initial_state(
            tokens.shape[0],
            tokens.device,
            tokens.dtype,
        )
        hidden = torch.zeros(
            tokens.shape[0],
            D_MODEL,
            device=tokens.device,
            dtype=tokens.dtype,
        )
        for step_index in range(tokens.shape[1]):
            hidden, states = self.step_token(tokens[:, step_index, :], states)
        return hidden, states

    def predict_from_hidden(
        self,
        hidden: torch.Tensor,
        previous_output: torch.Tensor,
    ) -> torch.Tensor:
        predicted_delta = DELTA_OUTPUT_SCALE * torch.tanh(self.head(hidden))
        return previous_output + predicted_delta

    def forward(
        self,
        input_history: torch.Tensor,
        output_history: torch.Tensor,
    ) -> torch.Tensor:
        hidden, _ = self.warmup(input_history, output_history)
        return self.predict_from_hidden(hidden, output_history[:, -1])

# -----------------------------------------------------------------------------
# 7. TRAINING
# -----------------------------------------------------------------------------

def count_trainable_parameters(model: nn.Module) -> int:
    return sum(parameter.numel() for parameter in model.parameters() if parameter.requires_grad)


def evaluate_one_step(
    model: nn.Module,
    loader: DataLoader,
    device: torch.device,
) -> float:
    model.eval()
    total_loss = 0.0
    total_samples = 0
    with torch.no_grad():
        for input_history, output_history, target in loader:
            input_history = input_history.to(device)
            output_history = output_history.to(device)
            target = target.to(device)
            prediction = model(input_history, output_history)
            loss = F.mse_loss(prediction, target)
            total_loss += float(loss.detach()) * len(target)
            total_samples += len(target)
    return total_loss / max(total_samples, 1)


def train_one_step(
    model: nn.Module,
    training_loader: DataLoader,
    validation_loader: DataLoader,
    device: torch.device,
    epochs: int,
) -> list[dict]:
    optimizer = torch.optim.AdamW(
        model.parameters(),
        lr=LEARNING_RATE,
        weight_decay=WEIGHT_DECAY,
    )
    history: list[dict] = []
    best_loss = float("inf")
    best_state: dict[str, torch.Tensor] | None = None

    for epoch in range(1, epochs + 1):
        model.train()
        total_loss = 0.0
        total_samples = 0

        for input_history, output_history, target in training_loader:
            input_history = input_history.to(device)
            output_history = output_history.to(device)
            target = target.to(device)

            # Small feedback noise improves robustness to imperfect histories.
            noisy_output_history = output_history + 0.005 * torch.randn_like(output_history)
            prediction = model(input_history, noisy_output_history)
            loss = F.mse_loss(prediction, target)

            optimizer.zero_grad(set_to_none=True)
            loss.backward()
            nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            optimizer.step()

            total_loss += float(loss.detach()) * len(target)
            total_samples += len(target)

        training_loss = total_loss / max(total_samples, 1)
        validation_loss = evaluate_one_step(model, validation_loader, device)

        history.append(
            {
                "stage": "mamba_one_step",
                "epoch": epoch,
                "teacher_forcing_ratio": 1.0,
                "training_normalized_mse": training_loss,
                "validation_normalized_mse": validation_loss,
            }
        )

        if validation_loss < best_loss:
            best_loss = validation_loss
            best_state = {
                name: value.detach().cpu().clone()
                for name, value in model.state_dict().items()
            }

        print(
            f"Mamba one-step {epoch:02d}/{epochs} | "
            f"train={training_loss:.7f} | validation={validation_loss:.7f}"
        )

    if best_state is None:
        raise RuntimeError("Mamba one-step training failed.")
    model.load_state_dict(best_state)
    return history


def recursive_batch_prediction(
    model: MambaActuatorModel,
    input_sequence: torch.Tensor,
    output_sequence: torch.Tensor,
    horizon: int,
    teacher_forcing_ratio: float,
) -> tuple[torch.Tensor, torch.Tensor]:
    input_history = input_sequence[:, :WINDOW]
    output_history = output_sequence[:, :WINDOW]
    hidden, states = model.warmup(input_history, output_history)

    predictions: list[torch.Tensor] = []
    targets: list[torch.Tensor] = []
    prediction = model.predict_from_hidden(hidden, output_history[:, -1])

    for step in range(horizon):
        target = output_sequence[:, WINDOW + step]
        predictions.append(prediction)
        targets.append(target)

        if teacher_forcing_ratio <= 0.0:
            feedback = prediction
        else:
            use_measured = (
                torch.rand(len(prediction), 1, device=prediction.device)
                < teacher_forcing_ratio
            )
            feedback = torch.where(use_measured, target, prediction)

        token = torch.cat(
            [input_sequence[:, WINDOW + step], feedback],
            dim=-1,
        )
        hidden, states = model.step_token(token, states)
        prediction = model.predict_from_hidden(hidden, feedback)

    return torch.stack(predictions, dim=1), torch.stack(targets, dim=1)


def evaluate_rollout(
    model: nn.Module,
    loader: DataLoader,
    device: torch.device,
    horizon: int,
) -> float:
    model.eval()
    losses: list[float] = []
    with torch.no_grad():
        for input_sequence, output_sequence in loader:
            input_sequence = input_sequence.to(device)
            output_sequence = output_sequence.to(device)
            prediction, target = recursive_batch_prediction(
                model,
                input_sequence,
                output_sequence,
                horizon,
                teacher_forcing_ratio=0.0,
            )
            losses.append(float(F.mse_loss(prediction, target).detach()))
    return float(np.mean(losses)) if losses else float("inf")


def train_rollout(
    model: nn.Module,
    training_loader: DataLoader,
    validation_loader: DataLoader,
    device: torch.device,
    epochs: int,
    horizon: int,
) -> list[dict]:
    optimizer = torch.optim.AdamW(
        model.parameters(),
        lr=ROLLOUT_LEARNING_RATE,
        weight_decay=WEIGHT_DECAY,
    )
    history: list[dict] = []
    best_loss = float("inf")
    best_state: dict[str, torch.Tensor] | None = None

    for epoch in range(1, epochs + 1):
        # Reaches fully free-running feedback in the final epoch.
        teacher_forcing_ratio = max(0.0, 1.0 - epoch / max(epochs - 1, 1))
        if epoch == epochs:
            teacher_forcing_ratio = 0.0

        model.train()
        losses: list[float] = []

        for input_sequence, output_sequence in training_loader:
            input_sequence = input_sequence.to(device)
            output_sequence = output_sequence.to(device)
            prediction, target = recursive_batch_prediction(
                model,
                input_sequence,
                output_sequence,
                horizon,
                teacher_forcing_ratio,
            )
            loss = F.mse_loss(prediction, target)

            optimizer.zero_grad(set_to_none=True)
            loss.backward()
            nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            optimizer.step()
            losses.append(float(loss.detach()))

        training_loss = float(np.mean(losses)) if losses else float("inf")
        validation_loss = evaluate_rollout(model, validation_loader, device, horizon)

        history.append(
            {
                "stage": "mamba_rollout",
                "epoch": epoch,
                "teacher_forcing_ratio": teacher_forcing_ratio,
                "training_normalized_mse": training_loss,
                "validation_normalized_mse": validation_loss,
            }
        )

        if validation_loss < best_loss:
            best_loss = validation_loss
            best_state = {
                name: value.detach().cpu().clone()
                for name, value in model.state_dict().items()
            }

        print(
            f"Mamba rollout {epoch:02d}/{epochs} | "
            f"measured feedback={teacher_forcing_ratio:.2f} | "
            f"train={training_loss:.7f} | validation={validation_loss:.7f}"
        )

    if best_state is None:
        raise RuntimeError("Mamba rollout training failed.")
    model.load_state_dict(best_state)
    return history

# -----------------------------------------------------------------------------
# 8. TEST PREDICTIONS
# -----------------------------------------------------------------------------

def predict_one_step(
    model: MambaActuatorModel,
    normalized_input: np.ndarray,
    normalized_output: np.ndarray,
    device: torch.device,
) -> np.ndarray:
    """
    Streaming measured-feedback prediction.

    After the first WINDOW samples initialize the Mamba state, each measured
    output is supplied as feedback for the following one-step prediction. This
    is the Mamba analogue of series-parallel evaluation and avoids rebuilding
    every overlapping 120-sample window.
    """
    predicted = np.zeros_like(normalized_output)
    predicted[:WINDOW] = normalized_output[:WINDOW]

    model.eval()
    with torch.no_grad():
        input_history = torch.from_numpy(normalized_input[:WINDOW][None]).to(device)
        output_history = torch.from_numpy(normalized_output[:WINDOW][None]).to(device)
        hidden, states = model.warmup(input_history, output_history)
        prediction = model.predict_from_hidden(hidden, output_history[:, -1])

        for target_index in range(WINDOW, len(normalized_input)):
            predicted[target_index] = prediction.cpu().numpy()[0]
            token = torch.cat(
                [
                    torch.from_numpy(normalized_input[target_index][None]).to(device),
                    torch.from_numpy(normalized_output[target_index][None]).to(device),
                ],
                dim=-1,
            )
            hidden, states = model.step_token(token, states)
            measured_feedback = torch.from_numpy(
                normalized_output[target_index][None]
            ).to(device)
            prediction = model.predict_from_hidden(hidden, measured_feedback)

    return predicted


def predict_free_running(
    model: MambaActuatorModel,
    normalized_input: np.ndarray,
    normalized_output: np.ndarray,
    device: torch.device,
) -> np.ndarray:
    predicted = np.zeros_like(normalized_output)
    predicted[:WINDOW] = normalized_output[:WINDOW]

    model.eval()
    with torch.no_grad():
        input_history = torch.from_numpy(normalized_input[:WINDOW][None]).to(device)
        output_history = torch.from_numpy(normalized_output[:WINDOW][None]).to(device)
        hidden, states = model.warmup(input_history, output_history)
        prediction = model.predict_from_hidden(hidden, output_history[:, -1])

        for target_index in range(WINDOW, len(normalized_input)):
            predicted[target_index] = prediction.cpu().numpy()[0]
            token = torch.cat(
                [
                    torch.from_numpy(normalized_input[target_index][None]).to(device),
                    prediction,
                ],
                dim=-1,
            )
            hidden, states = model.step_token(token, states)
            prediction = model.predict_from_hidden(hidden, prediction)

    return predicted


def persistence_baseline(measured: np.ndarray) -> np.ndarray:
    baseline = measured.copy()
    baseline[WINDOW:] = measured[WINDOW - 1:-1]
    return baseline

# -----------------------------------------------------------------------------
# 9. METRICS AND VALIDITY
# -----------------------------------------------------------------------------

def output_validity(
    measured: np.ndarray,
    predicted: np.ndarray,
) -> tuple[bool, list[str]]:
    messages: list[str] = []
    valid = True
    definitions = [
        (0, "displacement", "mm", 30.0),
        (1, "force", "N", 1.0),
    ]

    for column, name, unit, minimum_limit in definitions:
        actual = measured[WINDOW:, column]
        estimate = predicted[WINDOW:, column]
        data_range = float(np.ptp(actual))
        limit = max(minimum_limit, 10.0 * data_range)
        maximum = float(np.nanmax(np.abs(estimate)))
        rmse = float(np.sqrt(np.nanmean((actual - estimate) ** 2)))
        if not np.isfinite(maximum) or maximum > limit:
            valid = False
            messages.append(
                f"{name} invalid: max |prediction|={maximum:.6g} {unit}, limit={limit:.6g} {unit}"
            )
        if not np.isfinite(rmse) or rmse > max(data_range, 1e-12):
            valid = False
            messages.append(
                f"{name} inaccurate rollout: RMSE={rmse:.6g} {unit} exceeds measured range={data_range:.6g} {unit}"
            )

    return valid, messages


def calculate_metrics(
    measured: np.ndarray,
    predicted: np.ndarray,
    evaluation_name: str,
    status: str = "valid",
) -> list[dict]:
    rows: list[dict] = []
    for column, output_name, unit in [
        (0, "Displacement", "mm"),
        (1, "Lorentz force", "N"),
    ]:
        actual = measured[:, column]
        estimate = predicted[:, column]
        error = actual - estimate
        mse = float(np.mean(error ** 2))
        rmse = float(np.sqrt(mse))
        mae = float(np.mean(np.abs(error)))
        denominator = float(np.sum((actual - actual.mean()) ** 2))
        r_squared = 1.0 - float(np.sum(error ** 2)) / denominator if denominator > 0 else float("nan")
        norm_denominator = float(np.linalg.norm(actual - actual.mean()))
        fit_percent = (
            100.0 * (1.0 - float(np.linalg.norm(error)) / norm_denominator)
            if norm_denominator > 0
            else float("nan")
        )
        rows.append(
            {
                "dataset": "147mA",
                "evaluation": evaluation_name,
                "output": output_name,
                "unit": unit,
                "status": status,
                "MSE": mse,
                "RMSE": rmse,
                "MAE": mae,
                "R2": r_squared,
                "fit_percent": fit_percent,
            }
        )
    return rows


def regression_statistics(actual: np.ndarray, estimate: np.ndarray) -> tuple[float, float, float]:
    mask = np.isfinite(actual) & np.isfinite(estimate)
    x = actual[mask]
    y = estimate[mask]
    if len(x) < 2 or float(np.std(x)) < 1e-14:
        return float("nan"), float("nan"), float("nan")
    slope, intercept = np.polyfit(x, y, 1)
    fitted = slope * x + intercept
    denominator = float(np.sum((y - y.mean()) ** 2))
    r2 = 1.0 - float(np.sum((y - fitted) ** 2)) / denominator if denominator > 0 else float("nan")
    return float(slope), float(intercept), float(r2)

# -----------------------------------------------------------------------------
# 10. OUTPUT HELPERS
# -----------------------------------------------------------------------------

def create_output_folders() -> tuple[Path, Path, Path]:
    """Create the same three-folder structure used by the LSTM simulation.

    After every run, the code creates:
        01_Report_Figures/<timestamp>
        02_Presentation_Figures/<timestamp>
        03_Complete_Results/<timestamp>
    
    The complete folder stores all CSV/model/JSON/figure outputs.
    The report and presentation folders receive copies of the PNG figures.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_folder = ROOT / "01_Report_Figures" / timestamp
    presentation_folder = ROOT / "02_Presentation_Figures" / timestamp
    complete_folder = ROOT / "03_Complete_Results" / timestamp

    for folder in (report_folder, presentation_folder, complete_folder):
        folder.mkdir(parents=True, exist_ok=False)

    return report_folder, presentation_folder, complete_folder


def copy_figures_for_report_and_presentation(
    complete_folder: Path,
    report_folder: Path,
    presentation_folder: Path,
) -> None:
    """Copy every generated PNG into the report and presentation folders."""
    for image_path in complete_folder.glob("*.png"):
        shutil.copy2(image_path, report_folder / image_path.name)
        shutil.copy2(image_path, presentation_folder / image_path.name)


def write_rows(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def save_figure(figure: plt.Figure, path: Path) -> None:
    figure.tight_layout()
    figure.savefig(path, dpi=220, bbox_inches="tight")
    plt.close(figure)

# -----------------------------------------------------------------------------
# 11. FIGURES
# -----------------------------------------------------------------------------

def make_split_figure(blocks: list[Block], raw_data: dict[str, np.ndarray], folder: Path) -> None:
    figure, axis = plt.subplots(figsize=(12, 5))
    sheets = DEVELOPMENT_SHEETS + [TEST_SHEET]
    positions = {sheet: index for index, sheet in enumerate(sheets)}
    styles = {
        "training": ("tab:blue", "Training"),
        "validation": ("tab:orange", "Validation"),
        "development_test": ("tab:purple", "Development test"),
        "gap": ("0.7", "Guard gap"),
    }
    shown: set[str] = set()

    for block in blocks:
        time_values = raw_data[block.sheet][:, 0]
        start_time = float(time_values[block.start])
        end_time = float(time_values[block.end - 1])
        color, label_text = styles[block.role]
        label = None if label_text in shown else label_text
        shown.add(label_text)
        axis.barh(
            positions[block.sheet],
            end_time - start_time,
            left=start_time,
            height=0.55,
            color=color,
            alpha=0.45 if block.role == "gap" else 1.0,
            label=label,
        )

    test_time = raw_data[TEST_SHEET][:, 0]
    axis.barh(
        positions[TEST_SHEET],
        float(test_time[-1] - test_time[0]),
        left=float(test_time[0]),
        height=0.55,
        color="tab:green",
        label="Independent test",
    )
    axis.set_yticks(list(positions.values()))
    axis.set_yticklabels([sheet.replace("DC_Offset_", "") for sheet in sheets])
    axis.set_xlabel("Time (s)")
    axis.set_ylabel("Dataset")
    axis.set_title("Mamba experiment: cyclic development split and unseen 147 mA test")
    axis.grid(True, axis="x", alpha=0.3)
    axis.legend()
    save_figure(figure, folder / "00_mamba_data_split.png")


def make_training_figure(history: list[dict], folder: Path) -> None:
    figure, axis = plt.subplots(figsize=(10, 5.5))
    for stage, label in [
        ("mamba_one_step", "One-step"),
        ("mamba_rollout", "Rollout fine-tuning"),
    ]:
        rows = [row for row in history if row["stage"] == stage]
        if not rows:
            continue
        epochs = np.arange(1, len(rows) + 1)
        axis.plot(epochs, [row["training_normalized_mse"] for row in rows], label=f"{label}: train")
        axis.plot(
            epochs,
            [row["validation_normalized_mse"] for row in rows],
            linestyle="--",
            label=f"{label}: validation",
        )
    axis.set_yscale("log")
    axis.set_xlabel("Epoch")
    axis.set_ylabel("Normalized MSE")
    axis.set_title("Mamba training and validation losses")
    axis.grid(True, which="both", alpha=0.3)
    axis.legend()
    save_figure(figure, folder / "01_mamba_training_history.png")


def make_tracking_figure(
    time_values: np.ndarray,
    measured: np.ndarray,
    one_step: np.ndarray,
    folder: Path,
) -> None:
    start = WINDOW
    zoom_start = max(start, int(0.75 * len(time_values)))
    figure, axes = plt.subplots(2, 2, figsize=(15, 8))

    axes[0, 0].plot(time_values[start:], measured[start:, 0], label="Measured", linewidth=1.8)
    axes[0, 0].plot(time_values[start:], one_step[start:, 0], "--", label="Mamba", linewidth=1.3)
    axes[0, 0].set_title("Complete displacement record")
    axes[0, 0].set_ylabel("Displacement (mm)")

    axes[1, 0].plot(time_values[start:], measured[start:, 1], label="Measured", linewidth=1.8)
    axes[1, 0].plot(time_values[start:], one_step[start:, 1], "--", label="Mamba", linewidth=1.3)
    axes[1, 0].set_title("Complete force record")
    axes[1, 0].set_ylabel("Lorentz force (N)")

    axes[0, 1].plot(time_values[zoom_start:], measured[zoom_start:, 0], label="Measured", linewidth=1.8)
    axes[0, 1].plot(time_values[zoom_start:], one_step[zoom_start:, 0], "--", label="Mamba", linewidth=1.3)
    axes[0, 1].set_title("High-frequency displacement zoom")
    axes[0, 1].set_ylabel("Displacement (mm)")

    axes[1, 1].plot(time_values[zoom_start:], measured[zoom_start:, 1], label="Measured", linewidth=1.8)
    axes[1, 1].plot(time_values[zoom_start:], one_step[zoom_start:, 1], "--", label="Mamba", linewidth=1.3)
    axes[1, 1].set_title("High-frequency force zoom")
    axes[1, 1].set_ylabel("Lorentz force (N)")

    for axis in axes.flat:
        axis.set_xlabel("Time (s)")
        axis.grid(True, alpha=0.3)
        axis.legend()

    figure.suptitle("147 mA unseen test: Mamba one-step measured versus predicted", fontsize=14)
    save_figure(figure, folder / "02_mamba_one_step_measured_vs_predicted.png")


def make_error_figure(
    time_values: np.ndarray,
    measured: np.ndarray,
    baseline: np.ndarray,
    one_step: np.ndarray,
    folder: Path,
) -> None:
    start = WINDOW
    figure, axes = plt.subplots(2, 1, figsize=(12, 7), sharex=True)

    axes[0].plot(time_values[start:], measured[start:, 0] - baseline[start:, 0], label="Persistence error")
    axes[0].plot(time_values[start:], measured[start:, 0] - one_step[start:, 0], label="Mamba error")
    axes[0].set_ylabel("Displacement error (mm)")
    axes[0].set_title("Displacement error")

    axes[1].plot(time_values[start:], measured[start:, 1] - baseline[start:, 1], label="Persistence error")
    axes[1].plot(time_values[start:], measured[start:, 1] - one_step[start:, 1], label="Mamba error")
    axes[1].set_ylabel("Force error (N)")
    axes[1].set_xlabel("Time (s)")
    axes[1].set_title("Lorentz-force error")

    for axis in axes:
        axis.grid(True, alpha=0.3)
        axis.legend()

    figure.suptitle("147 mA unseen test: Mamba one-step error versus persistence baseline")
    save_figure(figure, folder / "03_mamba_one_step_errors.png")


def make_regression_figure(
    measured: np.ndarray,
    baseline: np.ndarray,
    one_step: np.ndarray,
    folder: Path,
) -> None:
    start = WINDOW
    actual = measured[start:]
    predictions = [("Persistence", baseline[start:]), ("Mamba one-step", one_step[start:])]
    outputs = [(0, "Displacement", "mm"), (1, "Lorentz force", "N")]
    figure, axes = plt.subplots(2, 2, figsize=(12, 9))
    stride = max(1, len(actual) // 5000)

    for column, (name, estimate) in enumerate(predictions):
        for row, (output_column, output_name, unit) in enumerate(outputs):
            axis = axes[row, column]
            x = actual[:, output_column]
            y = estimate[:, output_column]
            slope, intercept, r2 = regression_statistics(x, y)
            lower = float(min(x.min(), y.min()))
            upper = float(max(x.max(), y.max()))
            line = np.linspace(lower, upper, 200)
            axis.scatter(x[::stride], y[::stride], s=8, alpha=0.3)
            axis.plot(line, line, "--", label="Perfect prediction")
            if np.isfinite(slope):
                axis.plot(line, slope * line + intercept, ":", label="Regression")
            axis.set_xlabel(f"Measured ({unit})")
            axis.set_ylabel(f"Predicted ({unit})")
            axis.set_title(f"{name}: {output_name}\nRegression R²={r2:.5f}")
            axis.grid(True, alpha=0.3)
            axis.legend(fontsize=8)

    figure.suptitle("147 mA unseen test: Mamba regression comparison", fontsize=14)
    save_figure(figure, folder / "04_mamba_regression.png")


def make_accuracy_table(metric_rows: list[dict], folder: Path) -> None:
    table_rows: list[list[str]] = []
    for row in metric_rows:
        table_rows.append(
            [
                row["evaluation"],
                row["output"],
                row["status"],
                f"{row['RMSE']:.6g}",
                f"{row['MAE']:.6g}",
                f"{row['R2']:.5f}",
                f"{row['fit_percent']:.3f}",
            ]
        )

    figure, axis = plt.subplots(figsize=(12, 5.5))
    axis.axis("off")
    table = axis.table(
        cellText=table_rows,
        colLabels=["Evaluation", "Output", "Status", "RMSE", "MAE", "R²", "Fit (%)"],
        loc="center",
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.5)
    table.scale(1.0, 1.45)
    axis.set_title("147 mA unseen-test Mamba metrics")
    save_figure(figure, folder / "05_mamba_accuracy_table.png")


def make_rollout_diagnostic(
    time_values: np.ndarray,
    measured: np.ndarray,
    free_running: np.ndarray,
    valid: bool,
    validity_messages: list[str],
    folder: Path,
) -> None:
    start = WINDOW
    figure, axes = plt.subplots(2, 1, figsize=(12, 7), sharex=True)
    axes[0].plot(time_values[start:], measured[start:, 0], label="Measured", linewidth=1.8)
    axes[0].plot(time_values[start:], free_running[start:, 0], "--", label="Mamba free-running", linewidth=1.2)
    axes[0].set_ylabel("Displacement (mm)")
    axes[0].set_title("Free-running displacement")

    axes[1].plot(time_values[start:], measured[start:, 1], label="Measured", linewidth=1.8)
    axes[1].plot(time_values[start:], free_running[start:, 1], "--", label="Mamba free-running", linewidth=1.2)
    axes[1].set_ylabel("Lorentz force (N)")
    axes[1].set_xlabel("Time (s)")
    axes[1].set_title("Free-running Lorentz force")

    for axis in axes:
        axis.grid(True, alpha=0.3)
        axis.legend()

    status_text = "VALID PHYSICAL RANGE" if valid else "INVALID / DIVERGED"
    detail = "\n".join(validity_messages) if validity_messages else "No physical-range violation detected."
    figure.suptitle(f"Mamba free-running diagnostic: {status_text}\n{detail}", fontsize=12)
    save_figure(figure, folder / "06_mamba_free_running_diagnostic.png")


def _compatible_lstm_pair(folder: Path) -> tuple[Path, Path] | None:
    """Return the metrics/prediction pair in one LSTM result folder."""
    candidates = [
        (
            folder / "all_structures_metrics.csv",
            folder / "147mA_all_structures_predictions.csv",
        ),
        # Legacy names from the earlier reviewed package.
        (
            folder / "all_structures_metrics_revised.csv",
            folder / "147mA_all_structures_predictions_original.csv",
        ),
    ]
    for metrics_path, prediction_path in candidates:
        if metrics_path.exists() and prediction_path.exists():
            return metrics_path, prediction_path
    return None


def find_latest_lstm_result_pair() -> tuple[Path, Path]:
    """Locate newest complete LSTM result pair, including sibling folders."""
    if LSTM_RESULTS_FOLDER:
        requested = Path(LSTM_RESULTS_FOLDER).expanduser().resolve()
        direct = _compatible_lstm_pair(requested)
        if direct is not None:
            return direct
        roots = [requested]
    else:
        roots = [ROOT, ROOT.parent]
        if ROOT.parent.parent != ROOT.parent:
            roots.append(ROOT.parent.parent)

    pairs: list[tuple[Path, Path]] = []
    seen: set[Path] = set()
    for root in roots:
        if not root.exists():
            continue
        print(f"Searching for LSTM results under: {root}")
        for pattern in ("**/all_structures_metrics.csv", "**/all_structures_metrics_revised.csv"):
            try:
                for metrics_path in root.glob(pattern):
                    folder = metrics_path.parent.resolve()
                    if folder in seen:
                        continue
                    seen.add(folder)
                    pair = _compatible_lstm_pair(folder)
                    if pair is not None:
                        pairs.append(pair)
            except (OSError, PermissionError):
                continue

    if not pairs:
        raise FileNotFoundError(
            "No complete LSTM result set was found. The revised code searched the "
            "Mamba folder, its parent, and sibling folders. Run the LSTM code first, "
            "or set LSTM_RESULTS_FOLDER to the exact timestamped LSTM result folder."
        )

    return max(pairs, key=lambda pair: max(pair[0].stat().st_mtime, pair[1].stat().st_mtime))

def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as file:
        return list(csv.DictReader(file))


def read_numeric_csv(path: Path) -> dict[str, np.ndarray]:
    """Read a numeric CSV into a dictionary of float arrays."""
    rows = read_csv_rows(path)
    if not rows:
        raise ValueError(f"The CSV is empty: {path}")
    columns = list(rows[0].keys())
    result: dict[str, np.ndarray] = {}
    for column in columns:
        try:
            result[column] = np.asarray(
                [float(row[column]) for row in rows],
                dtype=np.float64,
            )
        except (TypeError, ValueError, KeyError) as error:
            raise ValueError(
                f"Column {column!r} in {path} is not a complete numeric column."
            ) from error
    return result


def align_to_reference_time(
    reference_time: np.ndarray,
    source_time: np.ndarray,
    values: np.ndarray,
) -> np.ndarray:
    """Align one LSTM result channel to the current Mamba test time vector."""
    if (
        len(reference_time) == len(source_time)
        and np.allclose(reference_time, source_time, rtol=1e-7, atol=1e-9)
    ):
        return values.astype(np.float64, copy=True)
    if np.any(np.diff(source_time) <= 0):
        # Earlier LSTM exports can contain exact duplicate time rows. Sort and
        # keep the first value at each time so old and new result folders are
        # both usable for comparison.
        order = np.argsort(source_time, kind="stable")
        source_time = source_time[order]
        values = values[order]
        source_time, unique_indices = np.unique(source_time, return_index=True)
        values = values[unique_indices]
        if np.any(np.diff(source_time) <= 0):
            raise ValueError("The LSTM prediction time vector cannot be made strictly increasing.")
    if reference_time[0] < source_time[0] or reference_time[-1] > source_time[-1]:
        raise ValueError(
            "The LSTM prediction time range does not cover the Mamba test record."
        )
    return np.interp(reference_time, source_time, values).astype(np.float64)


def metric_lookup(rows: list[dict], evaluation: str, output: str) -> dict:
    for row in rows:
        if row["evaluation"] == evaluation and row["output"] == output:
            return row
    raise KeyError(f"Missing metric row: evaluation={evaluation}, output={output}")


def parity_panel(
    axis: plt.Axes,
    measured_values: np.ndarray,
    predicted_values: np.ndarray,
    title: str,
    unit: str,
) -> None:
    mask = np.isfinite(measured_values) & np.isfinite(predicted_values)
    x = measured_values[mask]
    y = predicted_values[mask]
    stride = max(1, len(x) // COMPARISON_SCATTER_MAX_POINTS)
    lower = float(min(np.min(x), np.min(y)))
    upper = float(max(np.max(x), np.max(y)))
    pad = 0.05 * (upper - lower) if upper > lower else 1.0
    line = np.linspace(lower - pad, upper + pad, 200)
    slope, intercept, regression_r2 = regression_statistics(x, y)

    axis.scatter(x[::stride], y[::stride], s=8, alpha=0.28)
    axis.plot(line, line, "--", linewidth=1.2, label="Perfect prediction")
    if np.isfinite(slope):
        axis.plot(
            line,
            slope * line + intercept,
            ":",
            linewidth=1.4,
            label="Regression",
        )
    axis.set_xlim(lower - pad, upper + pad)
    axis.set_ylim(lower - pad, upper + pad)
    axis.set_xlabel(f"Measured ({unit})")
    axis.set_ylabel(f"Predicted ({unit})")
    axis.set_title(f"{title}\nRegression $R^2$ = {regression_r2:.5f}")
    axis.grid(True, alpha=0.3)
    axis.legend(fontsize=7)


def make_lstm_mamba_comparison(
    time_values: np.ndarray,
    measured: np.ndarray,
    persistence_prediction: np.ndarray,
    mamba_one_step: np.ndarray,
    mamba_free_running: np.ndarray,
    folder: Path,
) -> list[str]:
    """Create every LSTM-versus-Mamba result needed by the report/slides."""
    metrics_path, predictions_path = find_latest_lstm_result_pair()
    print("Using LSTM metrics:", metrics_path)
    print("Using LSTM predictions:", predictions_path)

    lstm_csv = read_numeric_csv(predictions_path)
    required_columns = {
        "time_s",
        "measured_displacement_mm",
        "series_displacement_mm",
        "series_parallel_displacement_mm",
        "parallel_displacement_mm",
        "measured_force_N",
        "series_force_N",
        "series_parallel_force_N",
        "parallel_force_N",
    }
    missing_columns = sorted(required_columns.difference(lstm_csv))
    if missing_columns:
        raise KeyError(
            "The LSTM prediction CSV is missing required columns: "
            + ", ".join(missing_columns)
        )

    lstm_time = lstm_csv["time_s"]
    def aligned(name: str) -> np.ndarray:
        return align_to_reference_time(time_values, lstm_time, lstm_csv[name])

    lstm_measured = np.column_stack(
        [aligned("measured_displacement_mm"), aligned("measured_force_N")]
    )
    measured_difference = float(np.nanmax(np.abs(lstm_measured - measured)))
    if measured_difference > 1e-5:
        print(
            "WARNING: LSTM and Mamba measured test columns differ by up to "
            f"{measured_difference:.6g}. The current Mamba workbook values are "
            "used as the common reference."
        )

    lstm_series = np.column_stack(
        [aligned("series_displacement_mm"), aligned("series_force_N")]
    )
    lstm_series_parallel = np.column_stack(
        [
            aligned("series_parallel_displacement_mm"),
            aligned("series_parallel_force_N"),
        ]
    )
    lstm_parallel = np.column_stack(
        [aligned("parallel_displacement_mm"), aligned("parallel_force_N")]
    )

    comparison_metrics: list[dict] = []
    comparison_metrics += calculate_metrics(
        measured[WINDOW:],
        persistence_prediction[WINDOW:],
        "Persistence baseline",
    )
    comparison_metrics += calculate_metrics(
        measured[WINDOW:],
        lstm_series[WINDOW:],
        "LSTM series",
    )
    comparison_metrics += calculate_metrics(
        measured[WINDOW:],
        lstm_series_parallel[WINDOW:],
        "LSTM series-parallel",
    )
    comparison_metrics += calculate_metrics(
        measured[WINDOW:],
        mamba_one_step[WINDOW:],
        "Mamba one-step",
    )

    # Add ranks and relative differences separately for displacement and force.
    for output_name in ["Displacement", "Lorentz force"]:
        output_rows = [row for row in comparison_metrics if row["output"] == output_name]
        output_rows.sort(key=lambda row: float(row["RMSE"]))
        best_rmse = float(output_rows[0]["RMSE"])
        lstm_sp_rmse = float(
            metric_lookup(comparison_metrics, "LSTM series-parallel", output_name)["RMSE"]
        )
        for rank, row in enumerate(output_rows, start=1):
            row["rmse_rank"] = rank
            row["relative_RMSE_vs_best_percent"] = (
                100.0 * (float(row["RMSE"]) - best_rmse) / best_rmse
                if best_rmse > 0 else float("nan")
            )
            row["relative_RMSE_vs_LSTM_series_parallel_percent"] = (
                100.0 * (float(row["RMSE"]) - lstm_sp_rmse) / lstm_sp_rmse
                if lstm_sp_rmse > 0 else float("nan")
            )

    metrics_csv_name = "lstm_mamba_comparison_metrics.csv"
    write_rows(folder / metrics_csv_name, comparison_metrics)

    generated = [metrics_csv_name]
    evaluation_start = WINDOW
    zoom_start = max(evaluation_start, int(0.75 * len(time_values)))

    # 07: RMSE comparison. Separate panels preserve displacement/force units.
    models = [
        "Persistence baseline",
        "LSTM series",
        "LSTM series-parallel",
        "Mamba one-step",
    ]
    figure, axes = plt.subplots(1, 2, figsize=(13, 5.6))
    for axis, output_name, unit in [
        (axes[0], "Displacement", "mm"),
        (axes[1], "Lorentz force", "N"),
    ]:
        values = [
            float(metric_lookup(comparison_metrics, model, output_name)["RMSE"])
            for model in models
        ]
        bars = axis.bar(np.arange(len(models)), values)
        axis.set_yscale("log")
        axis.set_xticks(np.arange(len(models)))
        axis.set_xticklabels(
            ["Persistence", "LSTM\nseries", "LSTM\nseries-parallel", "Mamba\none-step"],
            rotation=0,
        )
        axis.set_ylabel(f"RMSE ({unit}, logarithmic scale)")
        axis.set_title(f"{output_name}: lower is better")
        axis.grid(True, axis="y", which="both", alpha=0.3)
        for bar, value in zip(bars, values):
            axis.text(
                bar.get_x() + bar.get_width() / 2,
                value,
                f"{value:.4g}",
                ha="center",
                va="bottom",
                fontsize=8,
            )
    figure.suptitle("Unseen 147 mA test: LSTM and Mamba RMSE comparison", fontsize=14)
    rmse_name = "07_lstm_mamba_rmse_comparison.png"
    save_figure(figure, folder / rmse_name)
    generated.append(rmse_name)

    # 08: Direct fair one-step tracking comparison.
    figure, axes = plt.subplots(2, 2, figsize=(15, 8))
    plot_definitions = [
        (axes[0, 0], evaluation_start, 0, "Complete displacement", "Displacement (mm)"),
        (axes[1, 0], evaluation_start, 1, "Complete Lorentz force", "Lorentz force (N)"),
        (axes[0, 1], zoom_start, 0, "High-frequency displacement zoom", "Displacement (mm)"),
        (axes[1, 1], zoom_start, 1, "High-frequency force zoom", "Lorentz force (N)"),
    ]
    for axis, start, column, title, ylabel in plot_definitions:
        axis.plot(time_values[start:], measured[start:, column], label="Measured", linewidth=1.8)
        axis.plot(
            time_values[start:],
            lstm_series_parallel[start:, column],
            "--",
            label="LSTM series-parallel",
            linewidth=1.3,
        )
        axis.plot(
            time_values[start:],
            mamba_one_step[start:, column],
            ":",
            label="Mamba one-step",
            linewidth=1.4,
        )
        axis.set_title(title)
        axis.set_xlabel("Time (s)")
        axis.set_ylabel(ylabel)
        axis.grid(True, alpha=0.3)
        axis.legend(fontsize=8)
    figure.suptitle(
        "Unseen 147 mA test: fair measured-feedback comparison",
        fontsize=14,
    )
    tracking_name = "08_lstm_mamba_tracking_comparison.png"
    save_figure(figure, folder / tracking_name)
    generated.append(tracking_name)

    # 09: Common-reference error comparison.
    figure, axes = plt.subplots(2, 2, figsize=(15, 8))
    error_defs = [
        (axes[0, 0], evaluation_start, 0, "Complete displacement error", "Error (mm)"),
        (axes[1, 0], evaluation_start, 1, "Complete force error", "Error (N)"),
        (axes[0, 1], zoom_start, 0, "High-frequency displacement error", "Error (mm)"),
        (axes[1, 1], zoom_start, 1, "High-frequency force error", "Error (N)"),
    ]
    for axis, start, column, title, ylabel in error_defs:
        axis.plot(
            time_values[start:],
            measured[start:, column] - persistence_prediction[start:, column],
            label="Persistence",
            linewidth=1.0,
            alpha=0.75,
        )
        axis.plot(
            time_values[start:],
            measured[start:, column] - lstm_series_parallel[start:, column],
            label="LSTM series-parallel",
            linewidth=1.2,
        )
        axis.plot(
            time_values[start:],
            measured[start:, column] - mamba_one_step[start:, column],
            label="Mamba one-step",
            linewidth=1.2,
        )
        axis.axhline(0.0, linewidth=0.8)
        axis.set_title(title)
        axis.set_xlabel("Time (s)")
        axis.set_ylabel(ylabel)
        axis.grid(True, alpha=0.3)
        axis.legend(fontsize=8)
    figure.suptitle("Unseen 147 mA test: prediction-error comparison", fontsize=14)
    error_name = "09_lstm_mamba_error_comparison.png"
    save_figure(figure, folder / error_name)
    generated.append(error_name)

    # 10: Regression comparison with identical measured references.
    figure, axes = plt.subplots(2, 2, figsize=(12.5, 10))
    parity_panel(
        axes[0, 0],
        measured[evaluation_start:, 0],
        lstm_series_parallel[evaluation_start:, 0],
        "LSTM series-parallel: displacement",
        "mm",
    )
    parity_panel(
        axes[0, 1],
        measured[evaluation_start:, 0],
        mamba_one_step[evaluation_start:, 0],
        "Mamba one-step: displacement",
        "mm",
    )
    parity_panel(
        axes[1, 0],
        measured[evaluation_start:, 1],
        lstm_series_parallel[evaluation_start:, 1],
        "LSTM series-parallel: force",
        "N",
    )
    parity_panel(
        axes[1, 1],
        measured[evaluation_start:, 1],
        mamba_one_step[evaluation_start:, 1],
        "Mamba one-step: force",
        "N",
    )
    figure.suptitle("Unseen 147 mA test: LSTM versus Mamba regression", fontsize=14)
    regression_name = "10_lstm_mamba_regression_comparison.png"
    save_figure(figure, folder / regression_name)
    generated.append(regression_name)

    # 11: Complete metric table for report and presentation.
    table_rows = []
    for model in models:
        for output_name in ["Displacement", "Lorentz force"]:
            row = metric_lookup(comparison_metrics, model, output_name)
            table_rows.append(
                [
                    model,
                    output_name,
                    f"{float(row['RMSE']):.6g}",
                    f"{float(row['MAE']):.6g}",
                    f"{float(row['R2']):.6f}",
                    f"{float(row['fit_percent']):.3f}",
                    str(row["rmse_rank"]),
                ]
            )
    figure, axis = plt.subplots(figsize=(13, 6.2))
    axis.axis("off")
    table = axis.table(
        cellText=table_rows,
        colLabels=["Model", "Output", "RMSE", "MAE", "$R^2$", "Fit (%)", "RMSE rank"],
        loc="center",
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.5)
    table.scale(1.0, 1.45)
    axis.set_title("Unseen 147 mA test: common LSTM and Mamba accuracy table")
    table_name = "11_lstm_mamba_accuracy_table.png"
    save_figure(figure, folder / table_name)
    generated.append(table_name)

    # 12: Long-horizon/free-running stability comparison.
    lstm_valid, lstm_messages = output_validity(measured, lstm_parallel)
    mamba_valid, mamba_messages = output_validity(measured, mamba_free_running)
    figure, axes = plt.subplots(2, 2, figsize=(15, 8))
    stability_defs = [
        (axes[0, 0], 0, lstm_parallel, "LSTM parallel displacement", "mm", lstm_valid),
        (axes[0, 1], 0, mamba_free_running, "Mamba free-running displacement", "mm", mamba_valid),
        (axes[1, 0], 1, lstm_parallel, "LSTM parallel force", "N", lstm_valid),
        (axes[1, 1], 1, mamba_free_running, "Mamba free-running force", "N", mamba_valid),
    ]
    for axis, column, prediction, title, unit, valid in stability_defs:
        axis.plot(time_values[evaluation_start:], measured[evaluation_start:, column], label="Measured")
        axis.plot(time_values[evaluation_start:], prediction[evaluation_start:, column], label="Free-running prediction")
        axis.set_title(f"{title}\nStatus: {'VALID' if valid else 'INVALID / DIVERGED'}")
        axis.set_xlabel("Time (s)")
        axis.set_ylabel(f"Output ({unit})")
        axis.grid(True, alpha=0.3)
        axis.legend(fontsize=8)
    detail_lines = [
        "LSTM: " + ("; ".join(lstm_messages) if lstm_messages else "within validity limits"),
        "Mamba: " + ("; ".join(mamba_messages) if mamba_messages else "within validity limits"),
    ]
    figure.suptitle(
        "Long-horizon recursive simulation comparison\n" + " | ".join(detail_lines),
        fontsize=11,
    )
    stability_name = "12_lstm_mamba_free_running_stability.png"
    save_figure(figure, folder / stability_name)
    generated.append(stability_name)

    # 13: Data-driven model-selection summary for the report/slides.
    lstm_disp = float(metric_lookup(comparison_metrics, "LSTM series-parallel", "Displacement")["RMSE"])
    mamba_disp = float(metric_lookup(comparison_metrics, "Mamba one-step", "Displacement")["RMSE"])
    lstm_force = float(metric_lookup(comparison_metrics, "LSTM series-parallel", "Lorentz force")["RMSE"])
    mamba_force = float(metric_lookup(comparison_metrics, "Mamba one-step", "Lorentz force")["RMSE"])
    disp_advantage = 100.0 * (mamba_disp - lstm_disp) / mamba_disp if mamba_disp > 0 else float("nan")
    force_advantage = 100.0 * (mamba_force - lstm_force) / mamba_force if mamba_force > 0 else float("nan")

    figure, axis = plt.subplots(figsize=(13, 7.2))
    axis.axis("off")
    summary = (
        "CURRENT MODEL-SELECTION RESULT\n\n"
        "Preferred valid one-step model: LSTM series-parallel\n\n"
        f"Displacement RMSE: LSTM = {lstm_disp:.6g} mm, Mamba = {mamba_disp:.6g} mm\n"
        f"LSTM reduction relative to Mamba: {disp_advantage:.1f}%\n\n"
        f"Force RMSE: LSTM = {lstm_force:.6g} N, Mamba = {mamba_force:.6g} N\n"
        f"LSTM reduction relative to Mamba: {force_advantage:.1f}%\n\n"
        "Interpretation for this project:\n"
        "• The dataset is small and the valid task is local one-step prediction with measured feedback.\n"
        "• The LSTM gating/recurrent bias fits this actuator dataset better in the current experiment.\n"
        "• The portable small Mamba model does not yet exploit its expected advantage on very long sequences.\n"
        "• Neither LSTM parallel nor Mamba free-running is currently acceptable for long-horizon simulation.\n\n"
        "This conclusion is specific to the present data split, hyperparameters, and implementations."
    )
    axis.text(
        0.03,
        0.97,
        summary,
        va="top",
        ha="left",
        fontsize=13,
        linespacing=1.35,
        bbox=dict(boxstyle="round,pad=0.8", facecolor="white", edgecolor="0.35"),
    )
    summary_name = "13_lstm_mamba_model_selection_summary.png"
    save_figure(figure, folder / summary_name)
    generated.append(summary_name)

    summary_json = {
        "lstm_metrics_source": str(metrics_path),
        "lstm_predictions_source": str(predictions_path),
        "preferred_valid_one_step_model": "LSTM series-parallel",
        "displacement_RMSE": {
            "LSTM_series_parallel_mm": lstm_disp,
            "Mamba_one_step_mm": mamba_disp,
            "LSTM_reduction_relative_to_Mamba_percent": disp_advantage,
        },
        "force_RMSE": {
            "LSTM_series_parallel_N": lstm_force,
            "Mamba_one_step_N": mamba_force,
            "LSTM_reduction_relative_to_Mamba_percent": force_advantage,
        },
        "LSTM_parallel_valid": lstm_valid,
        "Mamba_free_running_valid": mamba_valid,
        "interpretation": [
            "Current LSTM series-parallel result is better than current Mamba one-step result.",
            "The result is experiment-specific and does not imply that LSTM is universally better than Mamba.",
            "Both recursive free-running models require further stability work.",
        ],
    }
    summary_json_name = "lstm_mamba_comparison_summary.json"
    (folder / summary_json_name).write_text(
        json.dumps(summary_json, indent=2),
        encoding="utf-8",
    )
    generated.append(summary_json_name)

    manifest_name = "lstm_mamba_comparison_figure_manifest.txt"
    (folder / manifest_name).write_text(
        "Figures generated for the report and presentation:\n"
        + "\n".join(name for name in generated if name.lower().endswith(".png"))
        + "\n",
        encoding="utf-8",
    )
    generated.append(manifest_name)

    print("Created complete LSTM-versus-Mamba comparison outputs.")
    return generated


# -----------------------------------------------------------------------------
# 12. GITHUB AUTOMATIC PUSH
# -----------------------------------------------------------------------------

def run_git(command: str, working_folder: Path, check: bool = True) -> tuple[int, str]:
    """Run one Git command inside the project folder."""
    print(f"  $ {command}")
    result = subprocess.run(
        command,
        cwd=working_folder,
        shell=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    output = (result.stdout + result.stderr).strip()
    if output:
        print(output)
    if check and result.returncode != 0:
        raise RuntimeError(f"Git command failed:\n{command}\n\n{output}")
    return result.returncode, output


def git_push_to_github(repository_folder: Path) -> None:
    """Commit and push the code and generated results after a successful run."""
    print()
    print("=" * 78)
    print("Automatic GitHub push")
    print(f"Repository: {GIT_REPOSITORY_SSH}")
    print("=" * 78)

    repository_folder = Path(repository_folder).resolve()
    print(f"Git working folder: {repository_folder}")

    run_git("git --version", repository_folder)

    return_code, _ = run_git(
        "git rev-parse --is-inside-work-tree",
        repository_folder,
        check=False,
    )
    if return_code != 0:
        print("This folder is not a Git repository. Initializing Git now...")
        run_git("git init", repository_folder)
        run_git("git branch -M main", repository_folder, check=False)

    run_git(f'git config user.name "{GIT_USER_NAME}"', repository_folder)
    run_git(f'git config user.email "{GIT_USER_EMAIL}"', repository_folder)

    remote_code, current_remote = run_git(
        "git remote get-url origin",
        repository_folder,
        check=False,
    )
    if remote_code != 0:
        run_git(f"git remote add origin {GIT_REPOSITORY_SSH}", repository_folder)
    elif current_remote.strip() != GIT_REPOSITORY_SSH:
        run_git(f"git remote set-url origin {GIT_REPOSITORY_SSH}", repository_folder)
    else:
        print(f"Remote origin is already correct: {GIT_REPOSITORY_SSH}")

    run_git("git branch -M main", repository_folder, check=False)

    merge_head = repository_folder / ".git" / "MERGE_HEAD"
    if merge_head.exists():
        _, unresolved = run_git(
            "git diff --name-only --diff-filter=U",
            repository_folder,
            check=False,
        )
        if unresolved.strip():
            raise RuntimeError(
                "Git has an unfinished merge with unresolved conflicts. "
                "Resolve the conflicts and run the simulation again."
            )
        run_git("git add .", repository_folder)
        merge_code, merge_output = run_git(
            'git commit -m "Complete previous merge before automatic push"',
            repository_folder,
            check=False,
        )
        if merge_code != 0 and "nothing to commit" not in merge_output.lower():
            raise RuntimeError("Could not complete the previous Git merge:\n" + merge_output)

    run_git("git add .", repository_folder)
    run_git("git status", repository_folder, check=False)

    diff_code, _ = run_git("git diff --cached --quiet", repository_folder, check=False)
    if diff_code != 0:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        commit_message = f"Update Mamba simulation code and results - {timestamp}"
        commit_code, commit_output = run_git(
            f'git commit -m "{commit_message}"',
            repository_folder,
            check=False,
        )
        if commit_code != 0:
            raise RuntimeError("Git commit failed:\n" + commit_output)
        print("Commit completed successfully:", commit_message)
    else:
        print("No new local changes to commit.")

    pull_code, pull_output = run_git(
        "git pull origin main --allow-unrelated-histories --no-rebase --no-edit",
        repository_folder,
        check=False,
    )
    if pull_code != 0:
        raise RuntimeError(
            "Git pull failed. GitHub main has changes that need manual attention.\n\n"
            + pull_output
        )

    push_code, push_output = run_git(
        "git push -u origin main",
        repository_folder,
        check=False,
    )
    if push_code != 0:
        raise RuntimeError(
            "Git push failed. Check SSH with: ssh -T git@github.com\n\n" + push_output
        )

    print("Files pushed successfully to GitHub main branch.")
    print("=" * 78)

# -----------------------------------------------------------------------------
# 13. MAIN
# -----------------------------------------------------------------------------

def main() -> None:
    set_random_seed()
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    maximum_rows = 2500 if QUICK_MODE else None
    one_step_epochs = 1 if QUICK_MODE else ONE_STEP_EPOCHS
    rollout_epochs = 1 if QUICK_MODE else ROLLOUT_EPOCHS
    rollout_horizon = 8 if QUICK_MODE else ROLLOUT_HORIZON
    one_step_batch_size = 16 if QUICK_MODE else ONE_STEP_BATCH_SIZE
    rollout_batch_size = 2 if QUICK_MODE else ROLLOUT_BATCH_SIZE

    workbook_path = find_workbook()
    print("=" * 78)
    print("Mamba-only actuator system identification")
    print("=" * 78)
    print(f"Workbook: {workbook_path}")
    print(f"Device: {device}")
    print(f"Quick mode: {QUICK_MODE}")
    print("Development: 67, 87, 107, 127 mA")
    print("Independent test: 147 mA")

    workbook = open_workbook_safely(workbook_path)
    sheet_names = DEVELOPMENT_SHEETS + [TEST_SHEET]
    raw_data = {
        sheet: load_sheet(workbook, sheet, maximum_rows)
        for sheet in sheet_names
    }
    workbook.close()

    development_data = {sheet: raw_data[sheet] for sheet in DEVELOPMENT_SHEETS}
    blocks = create_cyclic_blocks(development_data, rollout_horizon)
    normalizer = create_normalizer(development_data, blocks)
    input_data, output_data = normalize_all_records(raw_data, normalizer)

    training_dataset = OneStepDataset(
        input_data,
        output_data,
        blocks,
        role="training",
        stride=TRAIN_STRIDE,
    )
    validation_dataset = OneStepDataset(
        input_data,
        output_data,
        blocks,
        role="validation",
        stride=VALIDATION_STRIDE,
    )
    rollout_training_dataset = RolloutDataset(
        input_data,
        output_data,
        blocks,
        role="training",
        stride=ROLLOUT_STRIDE,
        horizon=rollout_horizon,
    )
    rollout_validation_dataset = RolloutDataset(
        input_data,
        output_data,
        blocks,
        role="validation",
        stride=ROLLOUT_STRIDE,
        horizon=rollout_horizon,
    )

    if QUICK_MODE:
        training_dataset = Subset(training_dataset, range(min(32, len(training_dataset))))
        validation_dataset = Subset(validation_dataset, range(min(16, len(validation_dataset))))
        rollout_training_dataset = Subset(
            rollout_training_dataset, range(min(4, len(rollout_training_dataset)))
        )
        rollout_validation_dataset = Subset(
            rollout_validation_dataset, range(min(2, len(rollout_validation_dataset)))
        )

    if len(training_dataset) == 0 or len(validation_dataset) == 0:
        raise RuntimeError("The cyclic split produced no one-step training or validation windows.")
    if len(rollout_training_dataset) == 0 or len(rollout_validation_dataset) == 0:
        raise RuntimeError("The cyclic split produced no rollout sequences.")

    training_loader = DataLoader(
        training_dataset,
        batch_size=one_step_batch_size,
        shuffle=True,
    )
    validation_loader = DataLoader(
        validation_dataset,
        batch_size=one_step_batch_size,
        shuffle=False,
    )
    rollout_training_loader = DataLoader(
        rollout_training_dataset,
        batch_size=rollout_batch_size,
        shuffle=True,
    )
    rollout_validation_loader = DataLoader(
        rollout_validation_dataset,
        batch_size=rollout_batch_size,
        shuffle=False,
    )

    model = MambaActuatorModel().to(device)
    print(f"Trainable parameters: {count_trainable_parameters(model):,}")
    print(f"One-step training windows: {len(training_dataset):,}")
    print(f"One-step validation windows: {len(validation_dataset):,}")
    print(f"Rollout training sequences: {len(rollout_training_dataset):,}")

    start_time = time.perf_counter()
    history = train_one_step(
        model,
        training_loader,
        validation_loader,
        device,
        one_step_epochs,
    )

    one_step_model = copy.deepcopy(model).to(device)
    free_running_model = copy.deepcopy(model).to(device)
    history += train_rollout(
        free_running_model,
        rollout_training_loader,
        rollout_validation_loader,
        device,
        rollout_epochs,
        rollout_horizon,
    )
    training_seconds = time.perf_counter() - start_time

    test_input = input_data[TEST_SHEET]
    test_output = output_data[TEST_SHEET]
    one_step_normalized = predict_one_step(one_step_model, test_input, test_output, device)
    free_running_normalized = predict_free_running(
        free_running_model,
        test_input,
        test_output,
        device,
    )

    measured = measured_outputs(raw_data[TEST_SHEET])
    one_step_prediction = normalizer.restore_output(one_step_normalized)
    free_running_prediction = normalizer.restore_output(free_running_normalized)
    baseline_prediction = persistence_baseline(measured)

    free_running_valid, validity_messages = output_validity(
        measured,
        free_running_prediction,
    )
    if not free_running_valid:
        print("WARNING: Mamba free-running rollout is outside the physical validity limits.")
        for message in validity_messages:
            print("  -", message)

    metrics: list[dict] = []
    metrics += calculate_metrics(
        measured[WINDOW:],
        baseline_prediction[WINDOW:],
        "Persistence baseline",
    )
    metrics += calculate_metrics(
        measured[WINDOW:],
        one_step_prediction[WINDOW:],
        "Mamba one-step",
    )
    metrics += calculate_metrics(
        measured[WINDOW:],
        free_running_prediction[WINDOW:],
        "Mamba free-running",
        status="valid" if free_running_valid else "invalid_diverged",
    )

    report_folder, presentation_folder, complete_folder = create_output_folders()
    output_folder = complete_folder
    write_rows(output_folder / "mamba_training_history.csv", history)
    write_rows(output_folder / "mamba_metrics.csv", metrics)

    prediction_table = np.column_stack(
        [
            raw_data[TEST_SHEET][:, 0],
            measured[:, 0],
            baseline_prediction[:, 0],
            one_step_prediction[:, 0],
            free_running_prediction[:, 0],
            measured[:, 1],
            baseline_prediction[:, 1],
            one_step_prediction[:, 1],
            free_running_prediction[:, 1],
        ]
    )
    np.savetxt(
        output_folder / "147mA_mamba_predictions.csv",
        prediction_table,
        delimiter=",",
        comments="",
        header=(
            "time_s,measured_displacement_mm,persistence_displacement_mm,"
            "mamba_one_step_displacement_mm,mamba_free_running_displacement_mm,"
            "measured_force_N,persistence_force_N,mamba_one_step_force_N,"
            "mamba_free_running_force_N"
        ),
    )

    np.savez(
        output_folder / "mamba_normalizer_parameters.npz",
        input_mean=normalizer.input_mean,
        input_std=normalizer.input_std,
        output_mean=normalizer.output_mean,
        output_std=normalizer.output_std,
    )
    torch.save(one_step_model.state_dict(), output_folder / "mamba_one_step_model.pt")
    torch.save(free_running_model.state_dict(), output_folder / "mamba_free_running_model.pt")

    run_info = {
        "implementation": "portable_pure_pytorch_mamba1_selective_ssm",
        "official_fused_kernel_used": False,
        "device": str(device),
        "quick_mode": QUICK_MODE,
        "window": WINDOW,
        "d_model": D_MODEL,
        "d_state": D_STATE,
        "d_conv": D_CONV,
        "expand": EXPAND,
        "n_mamba_blocks": N_MAMBA_BLOCKS,
        "trainable_parameters": count_trainable_parameters(model),
        "one_step_epochs": one_step_epochs,
        "rollout_epochs": rollout_epochs,
        "rollout_horizon": rollout_horizon,
        "training_seconds": training_seconds,
        "free_running_valid": free_running_valid,
        "free_running_validity_messages": validity_messages,
        "development_sheets": DEVELOPMENT_SHEETS,
        "test_sheet": TEST_SHEET,
    }
    (output_folder / "mamba_run_information.json").write_text(
        json.dumps(run_info, indent=2),
        encoding="utf-8",
    )

    make_split_figure(blocks, raw_data, output_folder)
    make_training_figure(history, output_folder)
    make_tracking_figure(
        raw_data[TEST_SHEET][:, 0],
        measured,
        one_step_prediction,
        output_folder,
    )
    make_error_figure(
        raw_data[TEST_SHEET][:, 0],
        measured,
        baseline_prediction,
        one_step_prediction,
        output_folder,
    )
    make_regression_figure(
        measured,
        baseline_prediction,
        one_step_prediction,
        output_folder,
    )
    make_accuracy_table(metrics, output_folder)
    make_rollout_diagnostic(
        raw_data[TEST_SHEET][:, 0],
        measured,
        free_running_prediction,
        free_running_valid,
        validity_messages,
        output_folder,
    )
    try:
        comparison_outputs = make_lstm_mamba_comparison(
            raw_data[TEST_SHEET][:, 0],
            measured,
            baseline_prediction,
            one_step_prediction,
            free_running_prediction,
            output_folder,
        )
        comparison_completed = True
    except FileNotFoundError as error:
        comparison_outputs = []
        comparison_completed = False
        warning_text = str(error)
        print("WARNING:", warning_text)
        (output_folder / "LSTM_COMPARISON_NOT_GENERATED.txt").write_text(
            warning_text + "\n\nMamba-only results were still generated successfully.",
            encoding="utf-8",
        )

    copy_figures_for_report_and_presentation(
        complete_folder,
        report_folder,
        presentation_folder,
    )

    expected_files = [
        "00_mamba_data_split.png",
        "01_mamba_training_history.png",
        "02_mamba_one_step_measured_vs_predicted.png",
        "03_mamba_one_step_errors.png",
        "04_mamba_regression.png",
        "05_mamba_accuracy_table.png",
        "06_mamba_free_running_diagnostic.png",
        "mamba_metrics.csv",
        "147mA_mamba_predictions.csv",
        "mamba_one_step_model.pt",
        "mamba_free_running_model.pt",
    ]
    expected_files.extend(comparison_outputs)
    missing = [name for name in expected_files if not (output_folder / name).exists()]
    if missing:
        raise RuntimeError("Missing expected Mamba outputs: " + ", ".join(missing))

    print()
    print("=" * 78)
    print("Mamba simulation finished successfully")
    print(f"Training time: {training_seconds:.2f} s")
    print(f"Report figures: {report_folder}")
    print(f"Presentation figures: {presentation_folder}")
    print(f"Complete results: {complete_folder}")
    print("LSTM-versus-Mamba comparison figures were copied to both figure folders." if comparison_completed else "LSTM comparison was skipped because no LSTM result folder was found.")
    print(f"Free-running status: {'VALID' if free_running_valid else 'INVALID / DIVERGED'}")
    print("=" * 78)

    if AUTO_GIT_PUSH:
        git_push_to_github(ROOT)
    else:
        print("Automatic GitHub push is disabled (AUTO_GIT_PUSH = False).")

    if os.name == "nt":
        try:
            os.startfile(str(report_folder))
            os.startfile(str(presentation_folder))
            os.startfile(str(complete_folder))
        except OSError:
            pass


if __name__ == "__main__":
    main()
